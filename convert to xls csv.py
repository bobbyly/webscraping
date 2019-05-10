'''
import pyperclip
import re
import openpyxl as ox
from openpyxl import Workbook


pasted_data = pyperclip.paste()
pattern = re.compile("\s")
result = pattern.sub(",", pasted_data)
wb1 = Workbook()
ws1 = wb1.create_sheet('testing')  # insert at the end (default)
c = ws1['A1']
c.value = result

wb1.save('PyKAT testing.xlsx')

with open(input_file,  newline='\n') as data:

'''

import csv
import openpyxl as ox
import os
import datetime

currentDT = datetime.datetime.now()

KAT_style = None
pathstring = r'C:\cifs\home\bly6\Python Scripts\PyKAT\PyKAT'+currentDT.strftime("%Y%m%d")

output_file = pathstring + "\\" + 'PyKAT Report ' + currentDT.strftime("%Y%m%d") + ".xls"
KWL_report_wb = ox.Workbook()

for filename in os.listdir(pathstring):
    KAT_style = filename.split(".txt")[0]
    KWL_report_wb.create_sheet(KAT_style, 0)  # insert at the end (default)
    KWL_report_ws = KWL_report_wb.worksheets[0]
    with open(pathstring + "\\" + filename,  newline='\n') as data:
        reader = csv.reader(data, delimiter='\t')
        for row in reader:
            KWL_report_ws.append(row)

KWL_report_wb.save(output_file)
print(output_file + " saved")


