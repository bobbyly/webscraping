# -*- coding: utf-8 -*-
"""
Created on Wed Jan 30 22:09:58 2019

@author: bly6
"""

import openpyxl as ox

filepath = 'H:\My Offline Files\Desktop\Sheets\d20 d36 advert check 0402.xlsx'
wb = ox.load_workbook(filepath, data_only=True)
ws = wb['apologies']
cell_range = ws['A2':'A13']
activeKC = "A"+X

for cell in cell_range:
    if cell[0].value == activeKC:
        print(cell[0].coordinate)
#        likekc = str(ws['G'+str(cell[0].row)].value)
#       print('like kc is: '+ likekc)
