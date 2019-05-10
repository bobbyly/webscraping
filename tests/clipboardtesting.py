'''
import pyperclip
import re
import openpyxl as ox
from openpyxl import Workbook


pasted_data = pyperclip.paste()
pattern = re.compile("\s")
result = pattern.sub(",", pasted_data)
wb1 = Workbook()
ws1 = wb1.create_sheet('testing')  # insert at the end (default)
c = ws1['A1']
c.value = result

wb1.save('PyKAT testing.xlsx')

with open(input_file,  newline='\n') as data:

'''

import csv
import openpyxl as ox
import pyperclip


pathstring = r'C:\cifs\home\bly6\Python Scripts\PyKAT\PyKAT20190327'

pasted_data = pyperclip.paste()

notepad_file = open(pathstring + ".txt", "w+")
notepad_file.write(pasted_data)
notepad_file.close()

input_file = pathstring + ".txt"
output_file = pathstring + ".xls"

wb = ox.Workbook()
wb.create_sheet('KAT_style')  # insert at the end (default)
ws = wb.active
#ws = wb.worksheets[0]

with open(input_file,  newline='\n') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)

wb.save(output_file)


